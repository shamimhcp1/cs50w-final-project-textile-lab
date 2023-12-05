import sys
import openpyxl
from datetime import date, datetime
import os
import comtypes.client
import concurrent.futures
import win32com.client


buyer_name: str = "WE"
print("Buyer name:", buyer_name)

def get_input(prompt):
    try:
        user_input = input(prompt + " : ").strip().upper()
        if user_input == '-1':
            sys.exit()
        return user_input
    except EOFError:
        sys.exit()

def get_valid_date_input(prompt):
    while True:
        date_str = get_input(prompt + " (dd-mm-yy)")
        if date_str == "":
            return date.today().strftime("%d-%m-%Y")  # Use today's date if blank input
        try:
            date_obj = datetime.strptime(date_str, "%d-%m-%y").date()
            return date_obj.strftime("%d-%m-%Y")
        except ValueError:
            print("Invalid date format. Please enter in the format dd-mm-yy.")

def convert_excel_to_pdf(excel_file_name):
    print("Creating PDF file...")
    # Get the current directory of the script
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Specify the input Excel file and output PDF file paths
    input_excel_file = os.path.join(script_dir, excel_file_name + '.xlsx')
    output_pdf_file = os.path.join(script_dir, excel_file_name + '.pdf')

    def create_pdf():
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            wb = excel.Workbooks.Open(input_excel_file)
            wb.ExportAsFixedFormat(0, output_pdf_file, Quality=0)
            wb.Close(False)
            excel.Quit()
            return "Success"
        except Exception as e:
            return str(e)

    result = create_pdf()
    if result == "Success":
        return "Success"
    else:
        return f"Error: {result}"

# tear result
def get_tear_result(wash_tear_warp, wash_tear_weft, dev_format):
    # 1) >340, 2) <340 None Stitch, 3) <340 Stitch
    if dev_format == 1:
        requirment_warp = float(20)
        requirment_weft = float(20)
    elif dev_format == 2 or dev_format == 3:
        requirment_warp = float(15)
        requirment_weft = float(15)

    tear_result = ""

    # check wash_tear_warp and wash_tear_weft grater then or equal to requirment
    if wash_tear_warp and wash_tear_weft:
        # check cross tear "over" value
        ct_warp = wash_tear_warp.endswith("CT")
        ct_weft = wash_tear_weft.endswith("CT")

        # cross tear warp
        if ct_warp:
            if not ct_weft:
                wash_tear_weft = float(wash_tear_weft)
                if wash_tear_weft >= requirment_weft:
                    tear_result = "Ok"
                else:
                    tear_result = "Not Ok"
                
        # cross tear weft
        elif ct_weft:
            if not ct_warp:
                wash_tear_warp = float(wash_tear_warp)
                if wash_tear_warp >= requirment_warp:
                    tear_result = "Ok"
                else:
                    tear_result = "Not Ok"

        # no cross tear in both warp and weft
        else:
            wash_tear_warp = float(wash_tear_warp)
            wash_tear_weft = float(wash_tear_weft)
            
            if wash_tear_warp >= requirment_warp and wash_tear_weft >= requirment_weft:
                tear_result = "Ok"
            else:
                tear_result = "Not Ok"

    return tear_result, wash_tear_warp, wash_tear_weft


def get_tensile_result(tensile_warp, tensile_weft, dev_format):
    # 1) >340, 2) <340 None Stitch, 3) <340 Stitch
    if dev_format == 1:
        requirment_warp = float(300)
        requirment_weft = float(300)
    if dev_format == 2:
        requirment_warp = float(230)
        requirment_weft = float(230)
    if dev_format == 3:
        requirment_warp = float(200)
        requirment_weft = float(200)
        
    tensile_result = ""
    if tensile_warp and tensile_weft:
        tensile_warp = float(tensile_warp)
        tensile_weft = float(tensile_weft)
        
        if tensile_warp >= requirment_warp and tensile_weft >= requirment_weft:
            tensile_result = "Ok"
        else:
            tensile_result = "Not Ok"
    else:
        tensile_result = "Small Size"
            
    return tensile_result, tensile_warp, tensile_weft

# rubbing result
def get_rubbing_result(dry_rubbing, wet_rubbing, color_shade):
    # 1) Light-Mid, 2) Dark
    if color_shade == 1:
        dry_req = ["3/4", "4", "4/5"]
        wet_req = ["3", "3/4", "4", "4/5"]
    if color_shade == 2:
        dry_req = ["3", "3/4", "4", "4/5"]
        wet_req = ["2/3", "3", "3/4", "4", "4/5"]
        
    rubbing_result = ""
    if dry_rubbing and wet_rubbing:

        # check dry rubbing
        if dry_rubbing in dry_req:
            rubbing_result = "Ok"
        else:
            rubbing_result = "Not Ok"

        # check wet rubbing
        if wet_rubbing in wet_req:
            rubbing_result = "Ok"
        else:
            rubbing_result = "Not Ok"
    else:
        rubbing_result = "N/A"

    return rubbing_result

# calculate final result
def get_final_result(tear_result, tensile_result, rubbing_result):
    if tear_result == "Ok" and tensile_result == "Ok" and rubbing_result == "Ok":
        final_result = "Result is Ok"
    elif tear_result == "Ok" and tensile_result == "Ok" and rubbing_result == "N/A":
        final_result = "Result is Ok"
    elif tear_result == "Ok" and tensile_result == "Small Size" and rubbing_result == "Ok":
        final_result = "Result is Ok"
    elif tear_result == "Ok" and tensile_result == "Small Size" and rubbing_result == "N/A":
        final_result = "Result is Ok"
    else:
        final_result = "Result Not Ok"

    return final_result

last_receive_date = ""
while True:
    # ask user to select dev format
    while True:
        try:
            dev_format = int(input("choose dev format: 1) >340, 2) <340 None Stitch, 3) <340 Stitch : ").strip())
            if dev_format == 1 or dev_format == 2 or dev_format == 3:
                break
            else:
                raise ValueError
        except ValueError:
            pass
        
    # ask user to select color shade    
    while True:
        try:
            color_shade = int(input("choose color shade: 1) Light-Mid 2) Dark : ").strip())
            if color_shade == 1 or color_shade == 2:
                break
            else:
                raise ValueError
        except ValueError:
            pass

    # ask for develoment details
    style = get_input("style")
    color = get_input("color")
    if not color:
        color = "?"
    fab_ref = get_input("fab_ref")
    fab_supp = get_input("fab_supp")
    if not fab_supp:
        fab_supp = "?"
    sample_type = get_input("sample_type")

    # use last receive date as new report receive date
    if last_receive_date:
        print("last receive_date: ", last_receive_date)
        change_receive_date = input("change receive date? (y) : ").strip().lower()
        if change_receive_date == "y":
            receive_date = get_valid_date_input("receive_date")
            last_receive_date = receive_date
        else:
            receive_date = last_receive_date
    else:
        receive_date = get_valid_date_input("receive_date")
        last_receive_date = receive_date
        
    report_date = get_valid_date_input("report_date")

    dry_rubbing = get_input("dry_rubbing")
    wet_rubbing = get_input("wet_rubbing")

    raw_tear_warp = get_input("raw_tear_warp")
    raw_tear_weft = get_input("raw_tear_weft")

    wash_tear_warp = get_input("wash_tear_warp")
    wash_tear_weft = get_input("wash_tear_weft")

    tensile_warp = get_input("tensile_warp")
    tensile_weft = get_input("tensile_weft")

    # get wash tear result
    tear_result = get_tear_result(wash_tear_warp, wash_tear_weft, dev_format)

    # get tensile result
    tensile_result = get_tensile_result(tensile_warp, tensile_weft, dev_format)

    # get rubbing tear result
    rubbing_result = get_rubbing_result(dry_rubbing, wet_rubbing, color_shade)

    # calculate final result
    final_result = get_final_result(tear_result[0], tensile_result[0], rubbing_result)

    # Open the existing Excel file
    if dev_format == 1:
        existing_excel_file = "we_dev_format_340_over.xlsx"
    if dev_format == 2:
        existing_excel_file = "we_dev_format_340_below_none_stitch.xlsx"
    if dev_format == 3:
        existing_excel_file = "we_dev_format_340_below_stitch.xlsx"
        
    wb = openpyxl.load_workbook(existing_excel_file)
    ws = wb.active

    # Update the specific cells with new data
    ws["C10"] = style
    ws["C11"] = color
    ws["C12"] = sample_type
    ws["G8"] = receive_date
    ws["G9"] = report_date
    ws["G10"] = fab_ref
    ws["G11"] = fab_supp
    ws["D18"] = raw_tear_warp
    ws["E18"] = raw_tear_weft
    ws["D21"] = tear_result[1] # wash_tear_warp
    ws["E21"] = tear_result[2] # wash_tear_weft
    ws["D26"] = tensile_result[1] # tensile_warp
    ws["E26"] = tensile_result[2] # tensile_weft
    # 1) Light-Mid, 2) Dark
    if color_shade == 1:
        ws["D34"] = dry_rubbing
        ws["E34"] = wet_rubbing
        ws["D35"] = ""
        ws["E35"] = ""
        ws["G34"] = rubbing_result
        ws["G35"] = ""
    if color_shade == 2:
        ws["D34"] = ""
        ws["E34"] = ""
        ws["D35"] = dry_rubbing
        ws["E35"] = wet_rubbing
        ws["G34"] = ""
        ws["G35"] = rubbing_result
    
    ws["G21"] = tear_result[0]
    ws["G26"] = tensile_result[0]
    ws["C42"] = final_result

    # Save the updated Excel file with a new name
    if color == "?":
        excel_file_name = f"{style} - {sample_type} - Fabric Ref - {fab_ref} - ({report_date})"
    else:
        excel_file_name = f"{style} - {color} - {sample_type} - Fabric Ref - {fab_ref} - ({report_date})"
        
    wb.save(f"{excel_file_name}.xlsx")

    print(f"\nExcel file has been saved.")

    if convert_excel_to_pdf(excel_file_name):
        print(f"PDF file has been saved.")

    create_another_report = input("\nPress any key to create another report? (no): ").strip().lower()
    if create_another_report == 'no':
        break
