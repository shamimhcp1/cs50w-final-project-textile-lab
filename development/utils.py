import os
from openpyxl import load_workbook
from .models import DevReport
import win32com.client as win32
import pythoncom  # Import the pythoncom module


def convert_excel_to_pdf(excel_file_name, buyer_name):
    input_excel_file = os.path.abspath(os.path.join('development', 'static', 'development', 'reports', buyer_name, excel_file_name + '.xlsx'))
    output_pdf_file = os.path.join('development', 'static', 'development', 'reports', buyer_name, excel_file_name + '.pdf')

    if not os.path.exists(input_excel_file):
        return f"Error: Excel file not found at {input_excel_file}"

    try:
        # Initialize COM for the current thread
        pythoncom.CoInitialize()

        # Create an Excel application instance
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False  # Ensure Excel is not visible

        # Open the Excel file
        wb = excel.Workbooks.Open(input_excel_file)

        # Save the workbook as PDF
        wb.ExportAsFixedFormat(0, output_pdf_file, Quality=0)

        # Close the workbook and quit Excel
        wb.Close(True)
        excel.Quit()

        # Uninitialize COM for the current thread
        pythoncom.CoUninitialize()

        return "Success"
    except Exception as e:
        return f"Error: {str(e)}"


# tear result
def get_tear_result(wash_tear_warp, wash_tear_weft):
    requirment = float(20)
    tear_result = ""

    # check wash_tear_warp and wash_tear_weft grater then or equal to requirment
    if wash_tear_warp and wash_tear_weft:
        # check cross tear "ct" value
        ct_warp = wash_tear_warp.endswith("CT")
        ct_weft = wash_tear_weft.endswith("CT")

        # cross tear warp
        if ct_warp:
            if not ct_weft:
                wash_tear_weft = float(wash_tear_weft)
                if wash_tear_weft >= requirment:
                    tear_result = "Ok"
                else:
                    tear_result = "Not Ok"
                
        # cross tear weft
        elif ct_weft:
            if not ct_warp:
                wash_tear_warp = float(wash_tear_warp)
                if wash_tear_warp >= requirment:
                    tear_result = "Ok"
                else:
                    tear_result = "Not Ok"

        # no cross tear in both warp and weft
        else:
            wash_tear_warp = float(wash_tear_warp)
            wash_tear_weft = float(wash_tear_weft)
            
            if wash_tear_warp >= requirment and wash_tear_weft >= requirment:
                tear_result = "Ok"
            else:
                tear_result = "Not Ok"

    return tear_result, wash_tear_warp, wash_tear_weft


def get_tensile_result(tensile_warp, tensile_weft):
    requirment = float(250)
    tensile_result = ""
    if tensile_warp and tensile_weft:
        tensile_warp = float(tensile_warp)
        tensile_weft = float(tensile_weft)
        
        if tensile_warp >= requirment and tensile_weft >= requirment:
            tensile_result = "Ok"
        else:
            tensile_result = "Not Ok"
    else:
        tensile_result = "Small Size"
            
    return tensile_result, tensile_warp, tensile_weft

# rubbing result
def get_rubbing_result(dry_rubbing, wet_rubbing):
    rubbing_result = ""
    if dry_rubbing and wet_rubbing:
        dry_req = ["3", "3/4", "4", "4/5"]
        wet_req = ["2", "2/3", "3", "3/4", "4", "4/5"]
        rubbing_result = ""

        # check dry rubbing
        if dry_rubbing in dry_req:
            rubbing_result = "Ok"
        else:
            rubbing_result = "Not Ok"

        # check wet rubbing
        if wet_rubbing in wet_req:
            rubbing_result = "Ok"
        else:
            rubbing_result = "Not Ok"
    else:
        rubbing_result = "N/A"

    return rubbing_result

# calculate final result
def get_final_result(tear_result, tensile_result, rubbing_result):
    if tear_result == "Ok" and tensile_result == "Ok" and rubbing_result == "Ok":
        final_result = "Result is Ok"
    elif tear_result == "Ok" and tensile_result == "Ok" and rubbing_result == "N/A":
        final_result = "Result is Ok"
    elif tear_result == "Ok" and tensile_result == "Small Size" and rubbing_result == "Ok":
        final_result = "Result is Ok"
    elif tear_result == "Ok" and tensile_result == "Small Size" and rubbing_result == "N/A":
        final_result = "Result is Ok"
    else:
        final_result = "Result Not Ok"

    return final_result


def generate_report(dev_report):
    # get wash tear result
    tear_result = get_tear_result(dev_report.wash_tear_warp, dev_report.wash_tear_weft)

    # get tensile result
    tensile_result = get_tensile_result(dev_report.tensile_warp, dev_report.tensile_weft)

    # get rubbing tear result
    rubbing_result = get_rubbing_result(dev_report.dry_rubbing, dev_report.wet_rubbing)

    # calculate final result
    final_result = get_final_result(tear_result[0], tensile_result[0], rubbing_result)

    # Open the existing Excel file
    existing_excel_file = os.path.join('development', 'static', 'development', 'formats', "dev_format.xlsx")
    wb = load_workbook(existing_excel_file)
    ws = wb.active

    # Update the specific cells with new data
    ws["C7"] = dev_report.buyer.name
    ws["C9"] = dev_report.style
    ws["C10"] = dev_report.color
    ws["C11"] = dev_report.sample_type
    ws["G7"] = dev_report.receive_date
    ws["G8"] = dev_report.report_date
    ws["G9"] = dev_report.fab_ref
    ws["G10"] = dev_report.fab_supplier
    ws["D20"] = dev_report.raw_tear_warp
    ws["E20"] = dev_report.raw_tear_weft
    ws["D21"] = tear_result[1] # wash_tear_warp
    ws["E21"] = tear_result[2] # wash_tear_weft
    ws["D25"] = tensile_result[1] # tensile_warp
    ws["E25"] = tensile_result[2] # tensile_weft
    ws["D33"] = dev_report.dry_rubbing
    ws["E33"] = dev_report.wet_rubbing
    ws["G21"] = tear_result[0]
    ws["G25"] = tensile_result[0]
    ws["G33"] = rubbing_result
    ws["C40"] = final_result

    # Save the updated Excel file with a new name
    if dev_report.color == "?":
        excel_file_name = f"{dev_report.style} - {dev_report.sample_type} - Fabric Ref - {dev_report.fab_ref}"
    else:
        excel_file_name = f"{dev_report.style} - {dev_report.color} - {dev_report.sample_type} - Fabric Ref - {dev_report.fab_ref}"
            
    wb.save(os.path.join('development', 'static', 'development', 'reports', dev_report.buyer.name, excel_file_name + '.xlsx'))
    

    # Convert Excel to PDF
    message = convert_excel_to_pdf(excel_file_name, dev_report.buyer.name)
    return message
